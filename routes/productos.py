from flask import Blueprint, render_template
from flask import request, redirect, flash, session
from database.conexion import obtener_conexion
from utils.auth import login_required
from flask import jsonify
from flask_jwt_extended import jwt_required
from utils.auth import admin_required
from utils.logger import logger
import os

from flask import Blueprint, send_file
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import subprocess
from datetime import datetime


from werkzeug.utils import secure_filename

productos = Blueprint("productos", __name__)

# PAGINA PRINCIPAL
@productos.route("/")
# @login_required
def inicio():

    conexion = obtener_conexion()
    cursor = conexion.cursor()

    buscar = request.args.get("buscar", "")
    orden = request.args.get("orden", "")
    stock = request.args.get("stock", "")

    #CONSULTA BASE
    consulta = """
                SELECT * FROM Productos
                WHERE activo = 1
                AND nombre LIKE ?
            """
                                      
    #Filtrar stock Bajo
    if stock == "bajo":
        consulta += " AND cantidad <= 5"

    #Filtrar stock Alto
    elif stock == "alto":
        consulta += " AND cantidad > 5"

    #Filtar por precio
    if orden == "nombre_asc":
        consulta += " ORDER BY nombre ASC"

    elif orden == "precio_mayor":
        consulta += " ORDER BY precio DESC"
        
    elif orden == "precio_menor":
        consulta += " ORDER BY precio ASC"

    cursor.execute(consulta, (f"%{buscar}%",))
    productos = cursor.fetchall()


    #RESUMEN INVENTARIO

    cursor.execute("""
                SELECT
                   SUM(cantidad) AS total_unidades,
                   SUM(precio * cantidad) AS valor_total
                FROM Productos
                WHERE activo = 1
                """)
    
    resumen = cursor.fetchone()

    # Controlamos que si no hay productos, las variables no se rompan en None
    unidades_stock = resumen[0] if resumen[0] else 0
    valor_inventario = resumen[1] if resumen[1] else 0


    # 3. NUEVA ESTADÍSTICA: TOTAL PRODUCTOS ACTIVOS 
    cursor.execute("SELECT COUNT(*) FROM Productos WHERE activo = 1")
    total_productos_activos = cursor.fetchone()[0]


    # 4. NUEVA ESTADÍSTICA: TOTAL VENTAS REALIZADAS
    cursor.execute("SELECT SUM(cantidad) FROM Ventas")
    resultado_ventas = cursor.fetchone()[0]
    total_ventas = resultado_ventas if resultado_ventas else 0
    

    # === CONSULTAS PARA GRÁFICAS (DASHBOARD) ===

    # G1: Productos más vendidos (Top 5)
    cursor.execute("""
            SELECT TOP 5 p.nombre, SUM(v.cantidad) AS total_vendido
            FROM Ventas v
            JOIN Productos p ON v.producto_id = p.id
            GROUP BY p.nombre
            ORDER BY total_vendido DESC
         """)
    
    top_vendidos = cursor.fetchall()
        # Los separamos en listas simples para JavaScript. Usamos ABS() por si tus cantidades son negativas.
    g1_labels = [row[0] for row in top_vendidos]
    g1_data = [abs(row[1]) for row in top_vendidos]

        # G2: Productos con menor stock (Alertas de reabastecimiento - Top 5 más bajos)
    cursor.execute("""
            SELECT TOP 5 nombre, cantidad 
            FROM Productos 
            WHERE activo = 1 
            ORDER BY cantidad ASC
        """)
    menor_stock = cursor.fetchall()
    g2_labels = [row[0] for row in menor_stock]
    g2_data = [row[1] for row in menor_stock]

    
    cursor.execute("""
            SELECT TOP 5 nombre, (precio * cantidad) AS valor_inventario
            FROM Productos 
            WHERE activo = 1 
            ORDER BY valor_inventario DESC
        """)
    valor_prod = cursor.fetchall()
    g3_labels = [row[0] for row in valor_prod]
    g3_data = [float(row[1]) for row in valor_prod] # Convertimos a float para que JSON lo entienda sin problemas

        # G4: Ventas por día (Historial de los últimos 7 días con ventas)
        # Nota: Ajusta 'fecha_venta' si tu columna en la tabla Ventas se llama diferente (ej: 'fecha')
    cursor.execute("""
            SELECT TOP 7 CONVERT(VARCHAR, fecha, 103) AS fecha, SUM(cantidad) AS total
            FROM Ventas
            GROUP BY CONVERT(VARCHAR, fecha, 103)
            ORDER BY fecha DESC
        """)
    ventas_dias = cursor.fetchall()
        # Las invertimos para que cronológicamente se lean de izquierda a derecha (pasado a presente)
    g4_labels = [row[0] for row in ventas_dias][::-1]
    g4_data = [abs(row[1]) for row in ventas_dias][::-1]

    conexion.close()

    # Armamos un diccionario limpio para las estadísticas
    estadisticas = {
        "productos_activos": total_productos_activos,
        "unidades_stock": unidades_stock,
        "valor_inventario": valor_inventario,
        "total_ventas": total_ventas
    }

     # Consolidamos los datos de las gráficas
    datos_graficas = {
        "g1": {"labels": g1_labels, "data": g1_data},
        "g2": {"labels": g2_labels, "data": g2_data},
        "g3": {"labels": g3_labels, "data": g3_data},
        "g4": {"labels": g4_labels, "data": g4_data}
    }

    return render_template("index.html", 
                           productos=productos,
                           resumen=resumen,
                           estadisticas=estadisticas,
                           datos_graficas=datos_graficas) # <- Pasamos el diccionario a la plantilla
   
   
# GUARDAR PRODUCTO
@productos.route("/guardar-producto", methods=["POST"])
# @login_required
def guardar_producto():

    nombre = request.form["nombre"]
    precio = request.form["precio"]
    cantidad = request.form["cantidad"]
    descripcion = request.form["descripcion"]
    imagen = request.files["imagen"]

    # ELIMINAR ESPACIOS EN BLANCO AL INICIO Y FINAL DEL NOMBRE
    nombre = nombre.strip()
    descripcion = descripcion.strip()

    if imagen.filename == "":
        flash("⚠ Debes seleccionar una imagen.")
        return redirect("/")

    nombre_imagen = secure_filename(imagen.filename)

    extensiones_permitidas = ["jpg", "jpeg", "png"]
    extension = nombre_imagen.split(".")[-1].lower()

    if extension not in extensiones_permitidas:
        flash("⚠ Formato de imagen no permitido.")
        return redirect("/")


    ruta_imagen = os.path.join("static/imagenes", nombre_imagen)
    imagen.save(ruta_imagen)

   
    #VALIDACIÓN DE CAMPOS
    if nombre == "" or precio == "" or cantidad == "" or descripcion == "":
        flash("⚠ Todos los campos son obligatorios.")
        return redirect("/")
    
    if len(nombre) < 3:
        flash("⚠ El nombre debe tener al menos 3 caracteres.")
        return redirect("/")
    
    if len(descripcion) < 10:
        flash("⚠ La descripción debe tener al menos 10 caracteres.")
        return redirect("/")
    
    try:
        precio = float(precio)
        cantidad = int(cantidad)

    except ValueError:
        flash("⚠ Precio o cantidad invalidos.")
        return redirect("/")

    #CONVERTIR DATOS
    precio = float(precio)
    cantidad = int(cantidad)

    if precio <= 0:
        flash("⚠ El precio debe ser mayor a 0.")
        return redirect("/")
    
    if cantidad < 0:
        flash("⚠ La cantidad no puede ser negativa.")
        return redirect("/")
  

    conexion = obtener_conexion()

    cursor = conexion.cursor()

    cursor.execute("""
        SELECT * FROM Productos
        WHERE nombre = ?
    """, (nombre,))

    producto_existente = cursor.fetchone()

    if producto_existente:
        flash("⚠ Ya existe un producto con ese nombre.")
        conexion.close()

        return redirect("/")

    cursor.execute("""
        INSERT INTO Productos (nombre, precio, cantidad, descripcion, imagen)
        VALUES (?, ?, ?, ?, ?)
    """, (nombre, precio, cantidad, descripcion, nombre_imagen))

    conexion.commit()
    flash("✔Producto guardado correctamente.")

    conexion.close()
    
    return redirect("/")

@productos.route("/desactivar-producto/<int:id>", methods=["POST"])
@login_required
def desactivar_producto(id):

    conexion = obtener_conexion()


    cursor = conexion.cursor()

    cursor.execute("""
                        UPDATE Productos
                        SET activo = 0
                        WHERE Id = ?
                        """, (id,))

    conexion.commit()
    flash("Producto desactivado correctamente.")
    conexion.close()

    return redirect("/")


@productos.route("/papelera")
@login_required
def papelera():

    conexion = obtener_conexion()

    cursor = conexion.cursor()

    cursor.execute("SELECT * FROM Productos WHERE activo = 0")

    productos_desactivados = cursor.fetchall()

    conexion.close()

    return render_template("papelera.html", productos=productos_desactivados)



@productos.route("/reactivar-producto/<int:id>", methods=['POST'])
@login_required
def reactivar_producto(id):

    conexion = obtener_conexion()

    cursor = conexion.cursor()

    cursor.execute("""
                        UPDATE Productos
                        SET activo = 1
                        WHERE id = ?
                    """, (id,))

    conexion.commit()
    print("Producto reactivado correctamente.")

    conexion.close()

    return redirect("/papelera")


@productos.route("/eliminar-producto/<int:id>", methods=['POST'])
@login_required
def eliminar_producto(id):

    conexion = obtener_conexion()

    cursor = conexion.cursor()

    cursor.execute("""
                        DELETE FROM Productos
                        WHERE id = ?
                    """, (id,))

    conexion.commit()
    flash("✔Producto eliminado correctamente.")

    conexion.close()

    return redirect("/papelera")


@productos.route("/editar-producto/<int:id>")
@login_required
def editar_producto(id):

    conexion = obtener_conexion()

    cursor = conexion.cursor()

    cursor.execute("""
                SELECT * FROM Productos
                WHERE id = ?
            """, (id,)
            )
    
    producto = cursor.fetchone()
    conexion.close()
    

    return render_template(
                        "editar_producto.html",
                        producto=producto
                    )

@productos.route("/actualizar-producto/<int:id>", methods=['POST'])
@login_required
def actualizar_producto(id):

    nombre = request.form["nombre"]
    precio = float(request.form["precio"])
    cantidad = int(request.form["cantidad"])
    descripcion = request.form["descripcion"]
    
    conexion = obtener_conexion()

    cursor = conexion.cursor()

    cursor.execute("""
                        UPDATE Productos
                        SET nombre = ?, 
                            precio = ?, 
                            cantidad = ?, 
                            descripcion = ?
                        WHERE id = ?
                    """, (nombre, precio, cantidad, descripcion, id))
    
    conexion.commit()
    flash("Producto actualizado correctamente. ")

    return redirect("/")

@productos.route("/vender-producto/<int:id>", methods=["POST"])
@login_required
def vender_producto(id):

    try:

        cantidad_vendida = int(request.form["cantidad"])

        # VALIDAR CANTIDAD NEGATIVA
        if cantidad_vendida <= 0:

            flash("⚠ La cantidad vendida debe ser mayor a 0")
            return redirect("/")

        conexion = obtener_conexion()

        cursor = conexion.cursor()

        cursor.execute("""
            SELECT cantidad
            FROM Productos
            WHERE id = ?
        """, (id,))

        producto = cursor.fetchone()

        if not producto:

            flash("❌ Producto no encontrado")
            return redirect("/")

        stock_actual = producto[0]

        if cantidad_vendida > stock_actual:

            flash("⚠ Stock insuficiente")

            conexion.close()

            return redirect("/")

        nuevo_stock = stock_actual - cantidad_vendida

        cursor.execute("""
            UPDATE Productos
            SET cantidad = ?
            WHERE id = ?
        """, (nuevo_stock, id))

        cursor.execute("""
            INSERT INTO Ventas(producto_id, cantidad)
            VALUES(?, ?)
        """, (id, cantidad_vendida))

        conexion.commit()

        conexion.close()

        logger.info(
            f"Venta Realizada | Producto ID: {id} | Cantidad: {cantidad_vendida}"
        )

        flash("✅ Venta registrada")

    except Exception as error:

        logger.error(f"Error en la venta: {error}")

        flash("🚨 Ocurrió un error en la venta 🚨")

    return redirect("/")

# HISTORIAL DE VENTAS
@productos.route("/historial")
@login_required
def historial():

    conexion = obtener_conexion()

    cursor = conexion.cursor()

    cursor.execute("""
        SELECT
            V.id,
            P.nombre,
            V.cantidad,
            V.fecha
        FROM Ventas V
        INNER JOIN Productos P
            ON P.id = V.producto_id
        ORDER BY V.fecha DESC
    """)

    historial_ventas = cursor.fetchall()

    conexion.close()

    return render_template(
        "historial.html",
        historial=historial_ventas
    )

@productos.route("/ganancias")
@login_required
def ganancias():

    conexion = obtener_conexion()

    cursor = conexion.cursor()

    # HOY

    cursor.execute("""
        SELECT
            SUM(P.precio * V.cantidad)
        FROM Ventas V
        INNER JOIN Productos P
            ON P.id = V.producto_id
        WHERE CAST(V.fecha AS DATE) = CAST(GETDATE() AS DATE)
    """)

    hoy = cursor.fetchone()[0] or 0

    # MES

    cursor.execute("""
        SELECT
            SUM(P.precio * V.cantidad)
        FROM Ventas V
        INNER JOIN Productos P
            ON P.id = V.producto_id
        WHERE MONTH(V.fecha) = MONTH(GETDATE())
        AND YEAR(V.fecha) = YEAR(GETDATE())
    """)

    mes = cursor.fetchone()[0] or 0

    # AÑO

    cursor.execute("""
        SELECT
            SUM(P.precio * V.cantidad)
        FROM Ventas V
        INNER JOIN Productos P
            ON P.id = V.producto_id
        WHERE YEAR(V.fecha) = YEAR(GETDATE())
    """)

    anio = cursor.fetchone()[0] or 0

    conexion.close()

    return render_template(
        "ganancias.html",
        hoy=hoy,
        mes=mes,
        anio=anio
    )

@productos.route("/top-productos")
@login_required
def top_productos():

    conexion = obtener_conexion()

    cursor = conexion.cursor()

    cursor.execute("""
        SELECT
            P.nombre,
            SUM(V.cantidad) AS total_vendido
        FROM Ventas V
        INNER JOIN Productos P
            ON P.id = V.producto_id
        WHERE V.cantidad > 0
        GROUP BY P.nombre
        ORDER BY total_vendido DESC
    """)

    productos = cursor.fetchall()

    conexion.close()

    return render_template(
        "top_productos.html",
        productos=productos
    )

# API REST PARA OBTENER PRODUCTOS EN FORMATO JSON
@productos.route("/api/productos")
@admin_required
@jwt_required()
def api_productos():
        
        """
        Listar productos
        ---
        tags:
            - Productos

        security:
            - Bearer: []
        responses:
            200:
                description: Lista de productos
        """
            
        conexion = obtener_conexion()

        cursor = conexion.cursor()

        cursor.execute("""
                SELECT * FROM Productos
                WHERE activo = 1
             """)
        
        productos_db = cursor.fetchall()

        conexion.close()

        productos_json = []

        for producto in productos_db:

            productos_json.append({
                "id": producto[0],
                "nombre": producto[1],
                "precio": producto[2],
                "cantidad": producto[3],
                "descripcion": producto[4],
                "imagen": producto[7]
            })

        return jsonify(productos_json)


@productos.route("/api/user/productos")
@jwt_required()
def api_user_productos():

    conexion = obtener_conexion()

    cursor = conexion.cursor()

    cursor.execute("""
        SELECT nombre, precio, cantidad
        FROM Productos
        WHERE activo = 1
    """)

    productos_db = cursor.fetchall()

    conexion.close()

    productos_json = []

    for producto in productos_db:

        productos_json.append({
            "nombre": producto[0],
            "precio": producto[1],
            "cantidad": producto[2]
        })

    return jsonify(productos_json), 200

@productos.route("/api/productos", methods=["POST"])
@login_required
def crear_producto_api():

    try:

        datos = request.get_json()

        nombre = datos["nombre"]
        precio = datos["precio"]
        cantidad = datos["cantidad"]
        descripcion = datos["descripcion"]


        #VALIDACION DE CAMPOS

        if not nombre:
             return jsonify({
                  "success": False,
                  "mensaje": "El nombre es obligatorio"
             }), 400
        
        if not precio:
            return jsonify({
                    "success": False,
                    "mensaje": "El precio es obligatorio"
                }), 400
        
        if not cantidad:
             return jsonify({
                  "success": False,
                  "mensaje": "la cantidad es obligatoria"
             }), 400

        conexion = obtener_conexion()

        cursor = conexion.cursor()

        cursor.execute("""
            INSERT INTO Productos
            (nombre, precio, cantidad, descripcion)
            VALUES (?, ?, ?, ?)
        """, (nombre, precio, cantidad, descripcion))

        conexion.commit()

        conexion.close()

        return jsonify({
            "success": True,
            "mensaje": "Producto creado correctamente"
        }), 201
    
    except Exception as e:
         
        return jsonify({
            "success": False,
            "mensaje": str(e)
        }), 500

@productos.route("/api/productos/<int:id>", methods=["PUT"])
@login_required
def editar_producto_api(id):
    
        datos = request.get_json()

        nombre = datos["nombre"]
        precio = datos["precio"]
        cantidad = datos["cantidad"]
        descripcion = datos["descripcion"]

        conexion = obtener_conexion()

        cursor = conexion.cursor()

        cursor.execute("""
            UPDATE Productos
            SET nombre = ?, precio = ?, cantidad = ?, descripcion = ?
            WHERE id = ?
        """, (nombre, precio, cantidad, descripcion, id))

        conexion.commit()

        conexion.close()

        return jsonify({
            "mensaje": "Producto actualizado correctamente"
        }), 200

@productos.route("/api/productos/<int:id>", methods=["DELETE"])
@login_required
def eliminar_producto_api(id):
     
        conexion = obtener_conexion()

        cursor = conexion.cursor()

        cursor.execute("""
            DELETE FROM Productos
            WHERE id = ?
        """, (id,))

        conexion.commit()

        conexion.close()

        return jsonify({
            "mensaje": "Producto eliminado correctamente"
        }), 200

@productos.route("/backup")
@login_required
def crear_backup():
    try:
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"inventario_{fecha}.bak"
        
        # 1. Ruta temporal en el disco C accesible para SQL Server
        ruta_temporal_sql = f"C:\\temp_backups\\{nombre_archivo}"

        # 2. Comando usando Driver 17 (agregamos -C por seguridad)
        comando_lista = [
            "sqlcmd",
            "-S", "DESKTOP-Q681RM2\\SQLEXPRESS",
            "-Q", f"BACKUP DATABASE InventarioAVA TO DISK='{ruta_temporal_sql}'",
            "-C"
        ]

        resultado = subprocess.run(comando_lista, capture_output=True, text=True)
        
        if resultado.returncode != 0:
            raise Exception(resultado.stderr or resultado.stdout)

        # 3. Ruta final de tu proyecto donde tú quieres ver el archivo
        ruta_destino_proyecto = os.path.abspath(os.path.join("backups", nombre_archivo))

        # 4. Con la ayuda de Python, movemos el archivo de la carpeta temporal a tu app
        if os.path.exists(ruta_temporal_sql):
            import shutil
            shutil.move(ruta_temporal_sql, ruta_destino_proyecto)
            flash(f"✅ Respaldo creado correctamente en tu proyecto: {nombre_archivo}")
        else:
            raise Exception("El archivo no se encontró en la carpeta temporal.")

    except Exception as e:
        flash(f"❌ Error creando respaldo: {e}")

    return redirect("/")

@productos.route("/reporte/pdf")
@login_required
def generar_pdf():
    # 1. Obtener la información de la base de datos
    conexion = obtener_conexion()
    cursor = conexion.cursor()
    
    cursor.execute("""
        SELECT nombre, precio, cantidad, (precio * cantidad) AS total 
        FROM Productos 
        WHERE activo = 1 
        ORDER BY nombre ASC
    """)
    items = cursor.fetchall()
    
    # Obtener totales generales para el cierre del informe
    cursor.execute("SELECT SUM(cantidad), SUM(precio * cantidad) FROM Productos WHERE activo = 1")
    resumen = cursor.fetchone()
    total_unidades = resumen[0] if resumen[0] else 0
    total_valor = resumen[1] if resumen[1] else 0
    
    conexion.close()

    # 2. Configurar el archivo en memoria virtual (Evita llenar el disco de basura)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    # 3. Definir estilos visuales del reporte
    estilos = getSampleStyleSheet()
    
    estilo_titulo = ParagraphStyle(
        'TituloReporte',
        parent=estilos['Heading1'],
        fontSize=24,
        textColor=colors.HexColor("#393e46"),
        spaceAfter=12
    )
    
    estilo_texto = ParagraphStyle(
        'TextoNormal',
        parent=estilos['Normal'],
        fontSize=10,
        textColor=colors.HexColor("#555555"),
        spaceAfter=6
    )

    # Lista de elementos que construirán el PDF secuencialmente
    historia = []

    # Encabezado / Branding
    historia.append(Paragraph("📊 INFORME EJECUTIVO DE INVENTARIO", estilo_titulo))
    historia.append(Paragraph("Generado por: Sistema de Gestión Automatizado", estilo_texto))
    historia.append(Spacer(1, 15))

    # 4. Estructurar la Tabla de Datos
    # Definimos los títulos de las columnas
    datos_tabla = [["Producto", "Precio Unitario", "Existencias", "Valor Total Valorizado"]]
    
    # Insertamos los registros de la base de datos limpios
    for fila in items:
        datos_tabla.append([
            fila[0],                               # Nombre
            f"$ {fila[1]:,.2f}",                    # Precio
            str(fila[2]),                          # Cantidad
            f"$ {fila[3]:,.2f}"                     # Valorizado
        ])
        
    # Añadimos la fila de balance final de cierre
    datos_tabla.append(["TOTALES", "", str(total_unidades), f"$ {total_valor:,.2f}"])

    # Convertimos la estructura a un objeto de ReportLab
    tabla_reporte = Table(datos_tabla, colWidths=[200, 100, 80, 150])
    
    # Diseño estético de las celdas (Colores idénticos a tu paleta web)
    estilo_celdas = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#00adb5")), # Encabezado Turquesa
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'), # Valores numéricos a la derecha
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor("#f9f9f9")), # Celdas alternas grises limpias
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        # Estilo exclusivo para la fila de Totales (Última fila)
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#eef1f6")),
    ])
    tabla_reporte.setStyle(estilo_celdas)
    
    historia.append(tabla_reporte)

    # 5. Compilar y retornar el archivo ejecutable binario
    doc.build(historia)
    buffer.seek(0)
    
    return send_file(
        buffer, 
        as_attachment=True, 
        download_name="reporte_inventario.pdf", 
        mimetype="application/pdf"
    )


@productos.route("/reporte/excel")
@login_required
def generar_excel():
    # 1. Extraer los datos actualizados de SQL Server
    conexion = obtener_conexion()
    cursor = conexion.cursor()
    
    cursor.execute("""
        SELECT nombre, precio, cantidad, (precio * cantidad) AS total 
        FROM Productos 
        WHERE activo = 1 
        ORDER BY nombre ASC
    """)
    items = cursor.fetchall()
    conexion.close()

    # 2. Crear el libro y la hoja de trabajo en memoria
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Actualizado"
    
    # Asegurar que las líneas de cuadrícula normales de Excel sean visibles
    ws.views.sheetView[0].showGridLines = True

    # 3. Diseñar los Estilos visuales del documento (Combinando con tu Paleta Turquesa)
    fuente_titulo = Font(name="Calibri", size=16, bold=True, color="333333")
    fuente_cabecera = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fuente_datos = Font(name="Calibri", size=11, color="000000")
    fuente_totales = Font(name="Calibri", size=11, bold=True, color="000000")
    
    # Relleno Turquesa corporativo para los títulos de columnas
    relleno_turquesa = PatternFill(start_color="00ADB5", end_color="00ADB5", fill_type="solid")
    relleno_totales = PatternFill(start_color="EEF1F6", end_color="EEF1F6", fill_type="solid")
    
    # Alineaciones
    alinear_izquierda = Alignment(horizontal="left", vertical="center")
    alinear_derecha = Alignment(horizontal="right", vertical="center")
    
    # Bordes finos y limpios
    borde_delgado = Side(border_style="thin", color="DDDDDD")
    cuadrilla = Border(left=borde_delgado, right=borde_delgado, top=borde_delgado, bottom=borde_delgado)
    
    borde_doble_abajo = Side(border_style="double", color="333333")
    borde_arriba = Side(border_style="thin", color="333333")
    linea_totales = Border(top=borde_arriba, bottom=borde_doble_abajo)

    # 4. Construir el Contenido del documento
    # Título Principal
    ws["A1"] = "REPORTE ANALÍTICO DE INVENTARIO Y STOCK"
    ws["A1"].font = fuente_titulo
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 35
    
    # Dejar una fila en blanco por estética
    ws.row_dimensions[2].height = 15

    # Encabezados de Columnas (Fila 3)
    cabeceras = ["Descripción del Producto", "Precio Unitario", "Existencias en Stock", "Valor Total en Bodega"]
    ws.append(cabeceras)
    ws.row_dimensions[3].height = 25
    
    for col_idx in range(1, 5):
        celda = ws.cell(row=3, column=col_idx)
        celda.font = fuente_cabecera
        celda.fill = relleno_turquesa
        celda.alignment = alinear_izquierda if col_idx == 1 else alinear_derecha

    # 5. Volcar los registros e inyectar FÓRMULAS VIVAS de Excel
    fila_actual = 4
    for fila in items:
        ws.cell(row=fila_actual, column=1, value=fila[0]).alignment = alinear_izquierda  # Nombre
        ws.cell(row=fila_actual, column=2, value=float(fila[1])).number_format = '"$"#,##0.00' # Precio
        ws.cell(row=fila_actual, column=3, value=int(fila[2])).number_format = '#,##0'       # Cantidad
        
        # En lugar de pegar el total calculado estático, le metemos la fórmula matemática multiplicativa de Excel
        celda_total = ws.cell(row=fila_actual, column=4, value=f"=B{fila_actual}*C{fila_actual}")
        celda_total.number_format = '"$"#,##0.00'
        
        # Aplicar fuentes y bordes individuales a cada celda de datos
        for col_idx in range(1, 5):
            c = ws.cell(row=fila_actual, column=col_idx)
            c.font = fuente_datos
            c.border = cuadrilla
            if col_idx > 1:
                c.alignment = alinear_derecha
                
        ws.row_dimensions[fila_actual].height = 20
        fila_actual += 1

    # 6. Agregar Fila de Cierre con Fórmulas de Sumatoria Automática (`SUM`)
    ws.cell(row=fila_actual, column=1, value="TOTALES").alignment = alinear_izquierda
    ws.cell(row=fila_actual, column=3, value=f"=SUM(C4:C{fila_actual-1})").number_format = '#,##0'
    ws.cell(row=fila_actual, column=4, value=f"=SUM(D4:D{fila_actual-1})").number_format = '"$"#,##0.00'
    
    ws.row_dimensions[fila_actual].height = 24
    
    for col_idx in range(1, 5):
        c = ws.cell(row=fila_actual, column=col_idx)
        c.font = fuente_totales
        c.fill = relleno_totales
        c.border = linea_totales
        if col_idx > 1:
            c.alignment = alinear_derecha

    # 7. Auto-ajustar el ancho de las columnas dinámicamente según el texto más largo
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue # Ignoramos la celda combinada del título para no distorsionar el cálculo
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    # 8. Exportar binario final para descarga directa
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name="reporte_inventario.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )